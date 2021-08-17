# -*- coding: utf-8 -*-

#from types import SimpleNamespace
import datetime
import hashlib
import json
import logging
import logging.handlers
import openpyxl
import os
import pathlib
import random
import re
import sys
import time

class torolib(object):

    def __init__(self, redis):
        self.redis = redis


    def buildMessage(self, texts, textIndices, tags=[], variables={}, tagHistory=''):
        '''recursive function, determines the text which matches the tag, resolves possible further tags within the found text'''
        message = None
        textId = self.getTextId(texts = texts, textIndices=textIndices, tags=tags, variables=variables)
        if textId is not None:
            text = texts[textId]['text']
            # replace existing [tags] within the text with concrete texts fitting these tags
            regExp = re.search(r'(.*)\[([a-zA-Z0-9,\s]+)\](.*)',text)
            while regExp:
                newTag = regExp.group(2)
                newMessage = None
                if '|'+newTag+'|' not in tagHistory:
                    newTags = [tag.strip() for tag in newTag.split(',')]
                    newMessage = self.buildMessage(texts = texts, textIndices = textIndices, tags=newTags, variables=variables, tagHistory=tagHistory+'|'+newTag +'|')
                else:
                    logging.warning(f'a text with {newTag} leads to another nested text with {newTag}, processing stops here')
                if newMessage is None:
                    newMessage = {}
                text = regExp.group(1) + newMessage.get('text',f'({newTag})') + regExp.group(3)
                regExp = re.search(r'(.*)\[([a-zA-Z0-9,\s]+)\](.*)',text)
            # replace {variables} in the text with their respective value
            regExp = re.search(r'(.*)\{([a-zA-Z0-9,\s]+)\}(.*)',text)
            while regExp:
                variableName = regExp.group(2)
                if variableName in variables.keys():
                    text = regExp.group(1) + str(variables[variableName]) + regExp.group(3)
                else:
                    text = regExp.group(1) + variableName + regExp.group(3)
                regExp = re.search(r'(.*)\{([a-zA-Z0-9,\s]+)\}(.*)',text)
            message = {'text':text,'prio': texts[textId].get('prio',3)}
        return message



    def evaluateTextCondition(self, cond, variables={}):
        '''determines if a condition of a text applies, returns True or False'''
        #logging.debug(f'evaluateTextCondition was called: {cond} and {variables}') #todo löschen
        result = False
        reRes = re.match(r"^([A-Za-z]{1}[A-Za-z0-9_]+)\s*(==|!=|>=|<=|=|<|>)\s*(.*)$", cond)
        if reRes:
            varName = reRes.groups()[0]
            operator = reRes.groups()[1]
            value = reRes.groups()[2]
            if varName in variables.keys():
                if isinstance(variables[varName],int):
                    try:
                        value = int(value)
                    except ValueError:
                        logging.error(f"The value '{value}' for the variable '{varName}' cannot be converted to integer.")
                        value = -1
                elif isinstance(variables[varName],float):
                    try:
                        value = float(value)
                    except ValueError:
                        logging.error(f"The value '{value}' for the variable '{varName}' cannot be converted to integer.")
                        value = -1.0
                if isinstance(variables[varName],int):
                    try:
                        value = int(value)
                    except ValueError:
                        value = -1
                if (operator in ['=','=='] and variables[varName] == value) \
                    or (operator == '!=' and variables[varName] != value) \
                    or (operator == '<' and variables[varName] < value) \
                    or (operator == '>' and variables[varName] > value) \
                    or (operator == '>=' and variables[varName] >= value) \
                    or (operator == '<=' and variables[varName] <= value):
                    result = True
            else:
                logging.warning (f'The variable {varName} is not supported but used in a text')
        else:
            logging.warning(f'The condition {cond} is invalid and could not be interpreted.')
        return result


    def factoryReset(self, redis):
        '''deletes all redis keys and initializes the database with default keys and values'''
        redis.flushdb()
        redis.set('hasBeenUsed', 1)
        redis.hset('sensors','mockMotion', 0)
        redis.set('remoteSSH', 1)
        redis.set('volume',90)
        redis.set('talkingInitiative',1)
        redis.hset('sensors','motionNoLongerDetected',time.time())


    def getTextsFromExcel(self, fileName, dataStartRow=2, sheetName='texts'):
        '''reads an excel file as input for the text collection'''
        #toDo: there should be a validation of the input file here
        results = []
        wb = openpyxl.load_workbook(fileName)
        ws = wb.get_sheet_by_name(sheetName)
        for i in range(dataStartRow, ws.max_row+1):
            if ws.cell(row=i, column=1).value:
                result = {'text': ws.cell(row=i, column=1).value}
                tags = ws.cell(row=i, column=2).value
                if tags:
                    result['tags'] = [tag.strip() for tag in tags.split(',')]
                conds = ws.cell(row=i, column=3).value
                if conds:
                    result['conds'] = [cond.strip() for cond in conds.split(',')]
                prio = ws.cell(row=i, column=4).value
                if prio:
                    result['prio'] = prio
                cooldown = ws.cell(row=i, column=5).value
                if cooldown:
                    result['cooldown'] = cooldown
                results.append(result)
        return results


    def getJsonContent(self, fileName):
        '''reads the last saved configuration from the sd card or disk'''
        with open(fileName, 'r') as file:
            fileContent = file.read()
        #content = json.loads(fileContent, object_hook=lambda d: SimpleNamespace(**d))
        content = json.loads(fileContent)
        return content


    def getTextId (self, texts, textIndices, tags, variables=[]):
        '''picks a randomly chosen text id  which a) fits the tags, b) has not been said recently, c) fits the texts conditions and d) has the highest priority class of all the texts for which a), b) and c) apply.'''
        textCandidates = []
        for prio in ['prio1','prio2','prio3']:
            textCandidates2ndChoice = []
            if textCandidates == []:
                for n in textIndices.get(prio,[]): # list of all text index numbers of the current prio
                    tagMatch = True
                    # check if all the tags from the function call match the current text candidate n
                    for tag in tags:
                        if n not in textIndices.get(tag,[]):
                            tagMatch = False
                    if tagMatch:
                        # check if the conditions match
                        cond = texts[n].get('conds',None)
                        condResult = None
                        if cond is None: # if a text has no conditions attached to it --> just use it
                            condResult = True
                        elif type(cond) == list: # if a text has multiple conditions --> evaluate them all and use the text only if ALL of them return True
                            for condItem in cond:
                                if condResult is not None:
                                    condResult = condResult and self.evaluateTextCondition(condItem, variables)
                                else:
                                    condResult = self.evaluateTextCondition(condItem, variables)
                        elif type(cond) == str: # if a text has exactly one condition --> use it
                            condResult = self.evaluateTextCondition(cond, variables)
                        else: # the text has a condition - but it cannot be parsed
                            logging.warning (f'an invalid data type was provided for the text condition {cond}')
                        if condResult:
                            # last check: has the text already been read during the cooldown time (usually 1 day)? If yes, other texts would be chosen whenever possible.
                            textHash = hashlib.md5(texts[n].get('text','').encode('utf-8')).hexdigest()
                            textLastUsed = self.redis.hget ('history',textHash)
                            cooldownPeriod = texts[n].get('cooldown',86400) # if no cooldown time has been specified for this text, the default  (1 day = 86400 sec) is used
                            if (textLastUsed is None) or (float(textLastUsed)+cooldownPeriod < datetime.datetime.now().timestamp()):
                                textCandidates.append(n)
                            else:
                                # textCandidates2ndChoice contains the texts, which apply to the situation but have already been used during their cooldown period
                                textCandidates2ndChoice.append(n)
        if textCandidates!=[]:
            textIndex = random.choice(textCandidates)
            self.redis.hset ('history',hashlib.md5(texts[textIndex].get('text','').encode('utf-8')).hexdigest(), datetime.datetime.now().timestamp())
        elif textCandidates2ndChoice!=[]:
            textIndex = random.choice(textCandidates2ndChoice)
            self.redis.hset ('history',hashlib.md5(texts[textIndex].get('text','').encode('utf-8')).hexdigest(), datetime.datetime.now().timestamp())
        else:
            textIndex = None
        return textIndex


    def getTextIndices(self, texts):
        '''for any standard text list this function creates a dict with indices: tag-names and priorities are mapped to index numbers in the original texts list'''
        result = {}
        result['prio1'] = []
        result['prio2'] = []
        result['prio3'] = []
        for n  in range(0,len(texts)):
            if 'text' in texts[n].keys():
                if type(texts[n]['text'])==str:
                    if texts[n]['text']!='':
                        if texts[n].get('prio',3) in [1,2,3]:
                            result['prio'+str(texts[n].get('prio',3))].append(n)
                        if 'tags' in texts[n].keys():
                            tags = texts[n]['tags']
                            for tag in tags:
                                if tag not in result.keys():
                                    result[tag] = []
                                result[tag].append(n)
        return result


    def prepareLoggingConfiguration(self, config):
        '''creates a dict with a logging configuration that can be used as **kwargs for calling logging.basicConfig'''
        result = {}
        # step 1: chose log level
        level = config.get('logLevel','info').lower()
        if level=='critical':
            result['level'] = logging.CRITICAL
        elif level=='error':
            result['level'] = logging.ERROR
        elif level in ['warning','warn']:
            result['level'] = logging.WARNING
        elif level=='debug':
            result['level'] = logging.DEBUG
        elif level=='notset':
            result['level'] = logging.NOTSET
        else:
            result['level'] = logging.INFO
        # step 2: prepare log handlers
        handlers = []
        if config.get('streamHandlerEnabled',0)==1:
            logHandler = logging.StreamHandler()
            logHandler.setLevel(result['level'])
            handlers.append(logHandler)
        if config.get('fileHandlerEnabled',0)==1:
            logFileName = config.get('logFile',None)
            if logFileName:
                if not os.path.isfile(logFileName):
                    folderName = os.path.dirname(logFileName)
                    if not os.path.isdir(folderName):
                        pathlib.Path(folderName).mkdir(parents=True, exist_ok=True)
                    with open(logFileName, 'w') as fp:
                        pass
                if os.path.isfile(logFileName):
                    logHandler = logging.handlers.WatchedFileHandler(logFileName)
                    logHandler.setLevel(result['level'])
                    handlers.append(logHandler)
                else:
                    sys.exit('The config file specifies a logFile which does not exist and cannot be created.')
            else:
                sys.exit('The config file wants the "fileHandlerEnabled" but does not provide a "logFile".')
        if len(handlers) > 0:
            result['handlers'] = handlers
        # step 3: set up format
        result['format']="%(asctime)s [%(levelname)s, %(module)s] %(message)s"
        return result
